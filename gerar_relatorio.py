import os
import json
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import (
    BarChart, 
    LineChart, 
    PieChart, 
    DoughnutChart,
    Reference, 
    Series,
    layout
)

# Tenta importar módulos do projeto, fallback se rodar isolado
try:
    from src.config import CONFIG
    from src.infrastructure.custom_logging import logger
    from src.infrastructure.repositories import JsonRepository
except ImportError:
    import logging
    logger = logging.getLogger("ReportFallback")
    class ConfigMock:
        BASE_DIR = Path(".")
        OUTPUT_DIR = Path("data/processed_billing")
    CONFIG = ConfigMock()
    class JsonRepository:
        def __init__(self, base): pass
        def load_filter_set(self, f): return set()

class ReportGenerator:
    def __init__(self):
        self.repo = JsonRepository(CONFIG.BASE_DIR)
        self.manifesto_set = self.repo.load_filter_set("manifesto.json")

    # --- Helpers de Conversão ---
    def _parse_date(self, date_str: Any) -> Optional[date]:
        if not date_str or not isinstance(date_str, str): return None
        try: return datetime.strptime(date_str, "%d/%m/%Y").date()
        except ValueError: return None

    def _parse_currency(self, value: Any) -> float:
        try: return float(value)
        except: return 0.0

    def _extract_parcelas(self, lista_parcelas: dict):
        if not lista_parcelas: return None, 0.0, 0, ""
        parcelas = lista_parcelas.get('parcela', [])
        if isinstance(parcelas, dict): parcelas = [parcelas]
        if not parcelas: return None, 0.0, 0, ""

        p1 = parcelas[0]
        dt_venc = self._parse_date(p1.get('data_vencimento'))
        valor_p1 = self._parse_currency(p1.get('valor', 0))
        dias = int(p1.get('quantidade_dias', 0))
        qtd = len(parcelas)
        resumo = f"{qtd}x" if qtd > 1 else "À Vista"
        
        return dt_venc, valor_p1, dias, resumo

    def _flatten_object(self, obj: dict, parent_key: str = '', sep: str = '_') -> dict:
        """Achata dicionários aninhados recursivamente."""
        items = []
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(self._flatten_object(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Listas não são achatadas aqui, são tratadas no processamento principal
                continue 
            else:
                items.append((new_key, v))
        return dict(items)

    # --- Core Processing ---
    def process_latest(self):
        target_file = self._get_latest_file()
        if not target_file:
            logger.warning("⚠️ Nenhum arquivo JSON encontrado para gerar relatório.")
            return

        logger.info(f"📂 Processando Analytics: {target_file.name}")
        self._process_json_file(target_file)

    def _get_latest_file(self) -> Optional[Path]:
        if not CONFIG.OUTPUT_DIR.exists():
            logger.error(f"Diretório de saída não existe: {CONFIG.OUTPUT_DIR}")
            return None
            
        files = list(CONFIG.OUTPUT_DIR.glob("*.json"))
        if not files: return None
        files.sort(key=os.path.getmtime, reverse=True)
        return files[0]

    def _process_json_file(self, json_file: Path):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                orders = json.load(f)

            all_rows = []
            
            # Itera sobre cada pedido
            for key, content in orders.items():
                # Normaliza estrutura (alguns jsons podem ter root 'pedido_venda_produto')
                details = content.get('pedido_venda_produto', content)
                
                # --- Extração de Blocos Principais ---
                cab = details.get('cabecalho', {})
                infoc = details.get('infoCadastro', {})
                adic = details.get('informacoes_adicionais', {})
                total = details.get('total_pedido', {})
                parcelas_obj = details.get('lista_parcelas', {})
                nota = details.get('nota_fiscal', {})
                obs = details.get('observacoes', {})
                
                # Hash de Integridade (Novo campo importante)
                hash_int = details.get('hash_integridade', 'N/A')

                # --- Datas e Referência ---
                d_fat = self._parse_date(infoc.get('dFat'))
                d_inc = self._parse_date(infoc.get('dInc'))
                d_prev = self._parse_date(cab.get('data_previsao'))
                # Data de referência para o BI: Faturamento > Previsão > Inclusão
                data_ref = d_fat or d_prev or d_inc
                
                if not data_ref: continue # Ignora pedidos sem data mínima

                # --- Métricas Calculadas ---
                lead_time = (d_fat - d_inc).days if (d_fat and d_inc) else None
                h_inc_str = infoc.get('hInc', '00:00:00')
                hora_pico = int(h_inc_str.split(':')[0]) if h_inc_str else 0
                venc_p1, valor_p1, dias_p1, condicao = self._extract_parcelas(parcelas_obj)

                # --- Dados do Cabeçalho (Comuns a todos os itens) ---
                header_data = {
                    "_DATA_REF": data_ref,
                    "Mes_Ano": data_ref.strftime("%Y-%m"),
                    "Ano": data_ref.year,
                    
                    # Identificação
                    "Numero Pedido": cab.get('numero_pedido'),
                    "ID Pedido": int(cab.get('codigo_pedido', 0)),
                    "Cliente ID": int(cab.get('codigo_cliente', 0)),
                    
                    # Comercial
                    "Vendedor": adic.get('vendedor_nome', "N/D"),
                    "Categoria": adic.get('categoria_nome', "N/D"),
                    "Valor Total Pedido": self._parse_currency(total.get('valor_total_pedido', 0)),
                    "Situação": "Cancelado" if infoc.get('cancelado') == 'S' else ("Faturado" if infoc.get('faturado') == 'S' else "Aberto"),
                    "Status Real": infoc.get('status_real', 'N/D'), # Campo novo do domain
                    
                    # Inteligência
                    "Lead Time": lead_time,
                    "Hora Inclusão": hora_pico,
                    "Condição Pagto": condicao,
                    
                    # Fiscal & Auditoria
                    "NFe Numero": nota.get('nNF', ''),
                    "NFe Emissão": nota.get('dEmi', ''),
                    "Tem NFe?": "Sim" if nota.get('nNF') else "Não",
                    "Integridade": "OK" if hash_int != 'N/A' else "Pendente",
                    "Hash Validação": hash_int,
                    
                    # Datas
                    "Dt. Faturamento": d_fat,
                    "Dt. Inclusão": d_inc,
                    "User Faturam.": infoc.get('uFat'),
                    "Manifesto": "S" if str(cab.get('codigo_pedido')) in self.manifesto_set else "",
                    "Obs Venda": obs.get('obs_venda', '')
                }

                # --- Processamento dos Itens (Explosão de Linhas) ---
                # Se houver lista de itens ('det'), criamos uma linha para cada item
                itens = details.get('det', [])
                if not isinstance(itens, list): itens = []
                
                if itens:
                    for item in itens:
                        prod = item.get('produto', {})
                        # Dados específicos do item
                        item_row = {
                            "Tipo Linha": "Item",
                            "Cod Produto": prod.get('codigo_produto'),
                            "Descricao Produto": prod.get('descricao', ''),
                            "NCM": prod.get('ncm', ''),
                            "CFOP": prod.get('cfop', ''),
                            "Qtd": self._parse_currency(prod.get('quantidade', 0)),
                            "Valor Unit": self._parse_currency(prod.get('valor_unitario', 0)),
                            "Valor Total Item": self._parse_currency(prod.get('valor_total', 0)),
                        }
                        # Combina Cabeçalho + Item
                        all_rows.append({**header_data, **item_row})
                else:
                    # Se não tiver itens (ex: pedido resumo), cria uma linha apenas com cabeçalho
                    dummy_item = {
                        "Tipo Linha": "Capa",
                        "Cod Produto": "", "Descricao Produto": "SEM ITENS / RESUMO",
                        "Qtd": 0, "Valor Unit": 0, "Valor Total Item": 0
                    }
                    all_rows.append({**header_data, **dummy_item})

            if not all_rows: 
                logger.warning("Nenhum dado processável encontrado.")
                return

            # Cria DataFrame
            df = pd.DataFrame(all_rows)
            
            # Ordenação inteligente
            if not df.empty:
                df = df.sort_values(by=['_DATA_REF', 'Numero Pedido'], ascending=[True, True])
            
            self._generate_excel(df)

        except Exception as e:
            logger.error(f"❌ Erro crítico ao processar JSON: {e}", exc_info=True)
            raise e

    # --- DASHBOARD ENGINE (Mantida e Ajustada) ---
    
    def _style_card(self, ws, ref, title, value, subtext="", color="1F4E78"):
        """Cria um cartão de KPI visualmente atraente."""
        cell = ws[ref]
        cell.value = value
        cell.font = Font(bold=True, size=20, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        thin = Side(border_style="thin", color="CCCCCC")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        col = get_column_letter(cell.column)
        r_title = cell.row - 1
        t_cell = ws[f"{col}{r_title}"]
        t_cell.value = title.upper()
        t_cell.font = Font(bold=True, size=9, color="666666")
        t_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        if subtext:
            r_sub = cell.row + 1
            s_cell = ws[f"{col}{r_sub}"]
            s_cell.value = subtext
            s_cell.font = Font(italic=True, size=8, color="888888")
            s_cell.alignment = Alignment(horizontal="center", vertical="top")

    def _create_dashboard(self, writer, df: pd.DataFrame):
        wb = writer.book
        ws = wb.create_sheet("DASHBOARD 360", 0)
        ws.sheet_view.showGridLines = False
        
        ws["B1"] = f"DASHBOARD DE VENDAS & OPERAÇÕES - {datetime.now().year}"
        ws["B1"].font = Font(bold=True, size=18, color="1F4E78")

        # --- 1. CÁLCULO DE MÉTRICAS (Ajustado para não duplicar totais por item) ---
        # Como explodimos os itens, precisamos deduplicar para contar pedidos e somar total do pedido
        df_pedidos = df.drop_duplicates(subset=['ID Pedido'])
        
        faturados = df_pedidos[df_pedidos["Situação"] == "Faturado"]
        
        total_fat = df_pedidos["Valor Total Pedido"].sum()
        total_orders = len(df_pedidos)
        avg_ticket = total_fat / total_orders if total_orders else 0
        
        lead_times = faturados["Lead Time"].dropna()
        avg_lead_time = lead_times.mean() if not lead_times.empty else 0
        
        com_nf = len(df_pedidos[df_pedidos["Tem NFe?"] == "Sim"])
        perc_fiscal = (com_nf / total_orders * 100) if total_orders else 0

        # --- 2. POSICIONAMENTO DOS CARDS ---
        self._style_card(ws, "B4", "Faturamento Total", total_fat, color="2E7D32")
        ws["B4"].number_format = 'R$ #,##0.00'
        
        self._style_card(ws, "D4", "Total Pedidos", total_orders)
        
        self._style_card(ws, "F4", "Ticket Médio", avg_ticket)
        ws["F4"].number_format = 'R$ #,##0.00'
        
        self._style_card(ws, "H4", "Lead Time Médio", f"{avg_lead_time:.1f} dias", "Inclusão -> Fat.", color="E65100")
        
        self._style_card(ws, "J4", "Cobertura Fiscal", f"{perc_fiscal:.1f}%", "Pedidos com NFe", color="1565C0")

        # --- 3. DADOS PARA GRÁFICOS ---
        
        # A. Evolução (Baseada em Pedidos Únicos)
        df_time = df_pedidos.groupby("Mes_Ano")["Valor Total Pedido"].sum().reset_index().sort_values("Mes_Ano")
        self._write_chart_data(ws, "AA5", ["Período", "Faturamento"], df_time)
        
        # B. Top Produtos (Baseada em Itens - AQUI USAMOS O DF COMPLETO)
        # Filtra apenas linhas de item para não pegar lixo
        df_items_only = df[df["Tipo Linha"] == "Item"]
        if not df_items_only.empty:
            df_prod = df_items_only.groupby("Descricao Produto")["Valor Total Item"].sum().nlargest(5).sort_values(ascending=True).reset_index()
            self._write_chart_data(ws, "AD5", ["Produto", "Total Vendido"], df_prod)
        else:
            self._write_chart_data(ws, "AD5", ["Produto", "Total"], pd.DataFrame([["Sem Dados", 0]]))
        
        # C. Categorias (Baseada em Pedidos)
        df_cat = df_pedidos.groupby("Categoria")["Valor Total Pedido"].sum().nlargest(6).reset_index()
        self._write_chart_data(ws, "AG5", ["Categoria", "Total"], df_cat)
        
        # D. Pico Horário
        df_hour = df_pedidos.groupby("Hora Inclusão")["ID Pedido"].count().reset_index()
        self._write_chart_data(ws, "AJ5", ["Hora", "Qtd Pedidos"], df_hour)

        # --- 4. GERAÇÃO DOS GRÁFICOS ---
        
        # Gráfico 1: Linha Temporal
        c1 = LineChart()
        c1.title = "Tendência de Faturamento"
        c1.style = 13
        c1.y_axis.title = "Valor (R$)"
        data = Reference(ws, min_col=28, min_row=6, max_row=6+len(df_time)-1)
        cats = Reference(ws, min_col=27, min_row=6, max_row=6+len(df_time)-1)
        c1.add_data(data, titles_from_data=False)
        c1.set_categories(cats)
        c1.height = 10; c1.width = 18
        ws.add_chart(c1, "B8")

        # Gráfico 2: Top Produtos (Mudamos de Vendedor para Produto para aproveitar a granularidade)
        c2 = BarChart()
        c2.type = "bar"
        c2.title = "Top 5 Produtos Mais Vendidos"
        c2.style = 11
        len_prod = len(df_prod) if not df_items_only.empty else 1
        data = Reference(ws, min_col=31, min_row=6, max_row=6+len_prod-1)
        cats = Reference(ws, min_col=30, min_row=6, max_row=6+len_prod-1)
        c2.add_data(data, titles_from_data=False)
        c2.set_categories(cats)
        c2.height = 10; c2.width = 15
        ws.add_chart(c2, "K8")

        # Gráfico 3: Categorias
        c3 = DoughnutChart()
        c3.title = "Share por Categoria"
        c3.style = 26
        len_cat = len(df_cat)
        data = Reference(ws, min_col=34, min_row=6, max_row=6+len_cat-1)
        cats = Reference(ws, min_col=33, min_row=6, max_row=6+len_cat-1)
        c3.add_data(data, titles_from_data=False)
        c3.set_categories(cats)
        c3.height = 10; c3.width = 15
        ws.add_chart(c3, "B28")

        # Gráfico 4: Horário
        c4 = BarChart()
        c4.type = "col"
        c4.title = "Horário de Pico"
        c4.style = 4
        len_hour = len(df_hour)
        data = Reference(ws, min_col=37, min_row=6, max_row=6+len_hour-1)
        cats = Reference(ws, min_col=36, min_row=6, max_row=6+len_hour-1)
        c4.add_data(data, titles_from_data=False)
        c4.set_categories(cats)
        c4.height = 10; c4.width = 18
        ws.add_chart(c4, "K28")

    def _write_chart_data(self, ws, start_cell, headers, df):
        col_letter = "".join(filter(str.isalpha, start_cell))
        start_row = int("".join(filter(str.isdigit, start_cell)))
        
        # Converte letra para índice (A=1, Z=26, AA=27...)
        # Hack rápido para colunas duplas AA, AB, etc.
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        
        ws.cell(row=start_row, column=col_idx, value=headers[0])
        ws.cell(row=start_row, column=col_idx+1, value=headers[1])
        
        for i, row in df.iterrows():
            ws.cell(row=start_row+1+i, column=col_idx, value=row[0])
            ws.cell(row=start_row+1+i, column=col_idx+1, value=row[1])

    def _generate_excel(self, df: pd.DataFrame):
        output_file = CONFIG.BASE_DIR / "Relatorio_BI_Avancado.xlsx"
        logger.info(f"📊 Gerando BI Avançado em: {output_file}")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._create_dashboard(writer, df)
            
            # Aba de Dados Detalhada
            export_df = df.drop(columns=['_DATA_REF', 'Mes_Ano', 'Ano'])
            export_df.to_excel(writer, sheet_name="BASE_DETALHADA", index=False)
        
        self._apply_advanced_styles(output_file)

    def _apply_advanced_styles(self, filename: Path):
        wb = load_workbook(filename)
        currency_fmt = 'R$ #,##0.00'
        date_fmt = 'dd/mm/yyyy'

        if "BASE_DETALHADA" in wb.sheetnames:
            ws = wb["BASE_DETALHADA"]
            
            # Transforma em Tabela Excel Real
            if ws.max_row > 1:
                ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                tab = Table(displayName="TabTransacoes", ref=ref)
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                ws.add_table(tab)
            
            # Ajuste de Largura e Formatação
            for col in ws.columns:
                header = str(col[0].value or "")
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = 18 
                
                if any(x in header for x in ["Valor", "Preço", "Total"]):
                    for cell in col[1:]: cell.number_format = currency_fmt
                elif any(x in header for x in ["Dt.", "Emissão"]):
                    for cell in col[1:]: cell.number_format = date_fmt
        
        wb.save(filename)
        logger.info("✅ Dashboard de BI (Next Level) gerado com sucesso!")

if __name__ == "__main__":
    gen = ReportGenerator()
    gen.process_latest()